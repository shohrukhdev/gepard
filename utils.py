from functools import wraps
from pprint import pprint

from asgiref.sync import sync_to_async
from telegram import Update

from bot.models import TelegramUser

def get_user(handler):
    @wraps(handler)
    async def wrapper(update: Update, context, *args, **kwargs):
        user = update.effective_user
        user_data = {
            'telegram_id': user.id,
            'username': user.username,
            'first_name': user.first_name,
            'last_name': user.last_name,
        }

        try:
            user = await TelegramUser.objects.prefetch_related("territory").aget(telegram_id=user.id)
            user.username = user_data.get("username", "")
            if not user.is_updated:
                user.first_name = user_data.get("first_name", "")
                user.last_name = user_data.get("last_name", "")

            await user.asave()
        except Exception as e:
            print(e)
            user = await TelegramUser.objects.acreate(**user_data)

        return await handler(update, context, user, *args, **kwargs)

    return wrapper


async def get_solo(model):
    return await sync_to_async(model.get_solo)()


@sync_to_async
def update_or_create(model, params, defaults):
    obj, created = model.objects.update_or_create(**params, defaults=defaults)
    return obj, created


from openpyxl import load_workbook


def import_data():
    from bot.models import Product
    workbook = load_workbook(filename="dumb.xlsx")
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        id, name, size, price_uzs_c, price_uzs_a, price_uzs_e, price_uzs_b, price_uzs_d = row
        price_uzs_a = str(price_uzs_a).replace(',00', '').replace(' ', '')
        price_uzs_b = str(price_uzs_b).replace(',00', '').replace(' ', '')
        price_uzs_c = str(price_uzs_c).replace(',00', '').replace(' ', '')
        price_uzs_d = str(price_uzs_d).replace(',00', '').replace(' ', '')
        price_uzs_e = str(price_uzs_e).replace(',00', '').replace(' ', '')
        Product.objects.create(title=name, price_uzs_a=price_uzs_a, price_uzs_b=price_uzs_b, price_uzs_c=price_uzs_c,
                               price_uzs_d=price_uzs_d, price_uzs_e=price_uzs_e)


def import_client_data():
    from bot.models import Area
    area, created = Area.objects.get_or_create(name="Мирзо Улугбек")
    workbook = load_workbook(filename="clients.xlsx")
    sheet = workbook.active
    import random
    for row in sheet.iter_rows(min_row=2, values_only=True):
        agent_tg_id = random.randint(999, 9999999)

        client, tin, agent = row[:3]
        u = TelegramUser.objects.create(telegram_id=agent_tg_id, first_name=client, tin=tin)
        u.territory.add(area)
