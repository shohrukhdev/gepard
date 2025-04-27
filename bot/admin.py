from collections.abc import Callable, Sequence
from typing import Any
from django.contrib import admin
from django.http import HttpRequest
from bot.models import TelegramUser, Contact, Product, Order, OrderItem, CustomUser, Category, Area, ArchiveOrder
from solo.admin import SingletonModelAdmin
from django.db.models import Sum, F, FloatField
from django.db.models import Q
from django.utils.html import format_html
from import_export.admin import ImportExportModelAdmin
from django.http import HttpResponseRedirect
from django.contrib.auth.models import Group
from django.utils import timezone
from datetime import timedelta

admin.site.unregister(Group)


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ("title", "cover")
    search_fields = ("title",)
    list_display_links = list_display

    def get_search_results(self, request, queryset, search_term):
        if search_term:
            search_term = search_term.lower()  # Ensure the search term is lowercase
            queries = [
                Q(**{f"{field}__icontains": search_term}) |
                Q(**{f"{field}__iregex": f"(?i){search_term}"})  # Using regex for case-insensitive match
                for field in self.search_fields
            ]
            queryset = queryset.filter(*queries)
        return queryset, False


@admin.register(CustomUser)
class CustomUserAdmin(admin.ModelAdmin):
    list_display = ("id", "username", "first_name", "last_name", "role",)
    fields = ("username", "first_name", "last_name", "role")
    list_display_links = list_display

    def get_fields(self, request: HttpRequest, obj=None):
        if not obj:
            return ("username", "first_name", "last_name", "role", "password")
        return super().get_fields(request, obj)

    def has_delete_permission(self, request: HttpRequest, obj=None):
        if obj and obj.is_superuser and not request.user.is_superuser:
            return False
        return True


admin.site.register(Contact, SingletonModelAdmin)


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ("id", "title", "is_active", "price_uzs_a",
                    "price_uzs_b", "price_uzs_c", "price_uzs_d", "price_uzs_e", "amount")
    list_editable = ('is_active', "price_uzs_a", "price_uzs_b", "price_uzs_c", "price_uzs_d", "price_uzs_e", "amount")
    list_display_links = ("id", "title",)
    list_filter = ("category", "is_active")
    search_fields = ("title",)

    def get_search_results(self, request, queryset, search_term):
        if search_term:
            search_term = search_term.lower()  # Ensure the search term is lowercase
            queries = [
                Q(**{f"{field}__icontains": search_term}) |
                Q(**{f"{field}__iregex": f"(?i){search_term}"})  # Using regex for case-insensitive match
                for field in self.search_fields
            ]
            queryset = queryset.filter(*queries)
        return queryset, False


from bot.resources import UsersTableResourse, OrderResource


@admin.register(TelegramUser)
class TelegramUserAdmin(ImportExportModelAdmin):
    list_display = ("id", "first_name", "username", "tin", "is_agent", "phone", "category")
    list_display_links = ("id", "first_name", "username", "phone", "category")
    list_editable = ("is_agent",)
    list_filter = ("category", "is_agent", "is_active")
    resource_classes = (UsersTableResourse,)
    skip_export_form = True
    search_fields = ("tin", "first_name", "username", "phone")
    readonly_fields = ("telegram_id", "phone")

    def get_search_results(self, request, queryset, search_term):
        if search_term:
            query = Q()
            for field in self.search_fields:
                query |= Q(**{f"{field}__icontains": search_term})
            queryset = queryset.filter(query)
        return queryset, False

class OrderItemTabularInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    readonly_fields = ['get_real_qty']

    def get_fields(self, request, obj=None):
        return ['product_name', 'product_in_set', 'qty', 'set_amount', 'price_uzs', 'get_real_qty']

    def has_delete_permission(self, request, obj=None) -> bool:
        return False

    def has_change_permission(self, request: HttpRequest, obj=None) -> bool:
        return False

    def has_add_permission(self, *args):
        return False


import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from django.utils.timezone import localtime
from django.http import HttpResponse


def export_orders_to_excel(modeladmin, request, queryset):
    # Create a new workbook and get the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Define the headers, based on your HTML structure
    # headers = [
    #     "Номер накладной", "Поставщик", "Дата отргрузки", "Дата консигнации", "Доставщик",
    #     "Время заказа", "Торговый представитель", "Телефон представителя", "Контрагент",
    #     "Тип оплаты", "Адрес", "Телефон"
    # ]
    # ws.append(headers)
    ws.column_dimensions['A'].width = 20  # Номер накладной
    ws.column_dimensions['B'].width = 25  # Поставщик
    ws.column_dimensions['C'].width = 15  # Дата отгрузки
    ws.column_dimensions['D'].width = 25  # Дата консигнации
    ws.column_dimensions['E'].width = 20  # Доставщик
    ws.column_dimensions['F'].width = 18  # Время заказа
    ws.column_dimensions['G'].width = 25  # Торговый представитель
    ws.column_dimensions['H'].width = 18  # Телефон представителя
    ws.column_dimensions['I'].width = 20  # Контрагент
    ws.column_dimensions['J'].width = 15  # Тип оплаты
    ws.column_dimensions['K'].width = 25  # Адрес
    ws.column_dimensions['L'].width = 18  # Телефон

    # Add each order's data to the workbook
    row_count = 0
    for order in queryset:
        if row_count != 0:
            ws.append(["", "", "", "", "", ""])
            ws.append(["", "", "", "", "", ""])
            ws.append(["", "", "", "", "", ""])

        order_agent_first_name = order.agent.first_name if order.agent and order.agent.first_name else "Нету информации"
        order_agent_last_name = order.agent.last_name if order.agent and order.agent.last_name else "Нету информации"
        col1 = ["Номер накладной", f"#{str(order.id).zfill(6)}", "", "Tорговый представитель",
                f"{order_agent_first_name} {order_agent_last_name}"]
        col2 = ["Поставщик", "", "", "Телефон представителя:", order.agent.phone if order.agent else "-"]
        col3 = ["Дата отргрузки", "", "", "Контрагент:", f"{order.user.first_name} {order.user.last_name}"]
        col4 = ["Дата консигнации", "", "", "Тип оплаты", order.get_payment_type_display()]
        col5 = ["Доставщик", "", "", "Адрес", order.user.territory.first().name]
        col6 = ["Время заказа", str(order.created_at.strftime("%d.%m.%Y %H:%M:%S")), "", "Телефон",
                str(order.user.phone if order.user.phone else "")]
        col7 = ["Комментария", str(order.comment)]

        # Append each row
        ws.append(col1)
        ws.append(col2)
        ws.append(col3)
        ws.append(col4)
        ws.append(col5)
        ws.append(col6)
        ws.append(col7)
        ws.append(["", "", "", "", "", ""])
        ws.append(["", "", "", "", "", ""])

        for row in range(row_count or 1, row_count + 8):
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"D{row}"].font = Font(bold=True)

        item_headers = ["#", "Продукция", "Количество", "Количество в блоке", "цена", "Сумма"]
        ws.append(item_headers)
        row_count += 9

        total_sum = 0
        for index, item in enumerate(order.items.all(), start=1):
            # Calculate the total in UZS
            item_total = float(item.price_uzs) * float(item.qty)
            total_sum += int(item_total)

            # Format price and total with thousands separators
            price_uzs_formatted = "{:,.0f}".format(float(item.price_uzs)).replace(",", " ")
            item_total_formatted = "{:,.0f}".format(item_total).replace(",", " ")

            # Prepare the data to append
            item_data = [
                index,
                item.product_name,
                item.qty,
                item.product_in_set,
                price_uzs_formatted,
                item_total_formatted,
            ]

            # Append to the worksheet
            ws.append(item_data)

            for col_num in range(1, len(item_data) + 1):
                cell = ws.cell(row=ws.max_row, column=col_num)
                cell.alignment = Alignment(horizontal="right")

        items_count = order.items.all().count()
        row_count += items_count

        ws.merge_cells(f"A{row_count + 1}:E{row_count + 1}")
        ws.merge_cells(f"A{row_count + 2}:E{row_count + 2}")
        ws.merge_cells(f"A{row_count + 3}:E{row_count + 3}")

        ws[f"A{row_count + 1}"] = "Сумма без переоценки"
        ws[f"A{row_count + 2}"] = "Сумма переоценки"
        ws[f"A{row_count + 3}"] = "Сумма с учётом НДС"
        ws[f"A{row_count + 1}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"A{row_count + 2}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"A{row_count + 3}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"A{row_count + 1}"].font = Font(bold=True)
        ws[f"A{row_count + 2}"].font = Font(bold=True)
        ws[f"A{row_count + 3}"].font = Font(bold=True)
        ws[f"F{row_count + 1}"] = "{:,.0f}".format(total_sum).replace(",", " ")

        ws.append([])  # Blank row for separation

        fill = PatternFill(start_color="efefef", end_color="efefef", fill_type="solid")
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in ws[f"A{row_count - items_count}:F{row_count + 3}"]:
            for cell in row:
                cell.border = border

        for row in ws[f"A{row_count - items_count}:F{row_count - items_count}"]:
            for cell in row:
                cell.fill = fill
                cell.font = Font(bold=True)

        for row in ws[f"A{row_count + 1}:F{row_count + 3}"]:
            for cell in row:
                cell.fill = fill

        ws.merge_cells(f"A{row_count + 5}:C{row_count + 5}")
        ws.merge_cells(f"A{row_count + 7}:C{row_count + 7}")
        ws.merge_cells(f"A{row_count + 9}:C{row_count + 9}")

        ws.merge_cells(f"D{row_count + 5}:F{row_count + 5}")
        ws.merge_cells(f"D{row_count + 7}:F{row_count + 7}")
        ws.merge_cells(f"D{row_count + 9}:F{row_count + 9}")
        ws[f"A{row_count + 5}"] = "Провадец: ____________________"
        ws[f"A{row_count + 7}"] = "Торговый представитель: ____________________"
        ws[f"D{row_count + 5}"] = "Получатель: ____________________"

        row_count += 12

    # Set up the HTTP response with the generated Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="orders.xlsx"'
    wb.save(response)
    return response


export_orders_to_excel.short_description = "Накладная для заказа (Excel)"


def export_invoice_total_amount(modeladmin, request, queryset):
    # Create a new workbook and get the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Define the headers, based on your HTML structure
    ws.column_dimensions['A'].width = 20  # Полъзователи
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    # Add each order's data to the workbook
    col1 = ["Полъзователи", f"{', '.join([order.user.get_full_name() for order in queryset])}", "", ]
    ws.append(col1)
    ws.append(["", "", "", "", "", ""])

    fill = PatternFill(start_color="efefef", end_color="efefef", fill_type="solid")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Append each row
    item_headers = ["#", "Название", "Количество в блоке", "Блок", "Кол-во", "Сумма"]
    ws.append(item_headers)
    row_count = 3
    total_count = 0
    total_sum = 0
    count = 0
    for row in range(row_count or 1, row_count + 1):
        ws[f"A{row}"].fill = fill
        ws[f"B{row}"].fill = fill
        ws[f"C{row}"].fill = fill
        ws[f"D{row}"].fill = fill
        ws[f"E{row}"].fill = fill
        ws[f"F{row}"].fill = fill
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].font = Font(bold=True)
        ws[f"C{row}"].font = Font(bold=True)
        ws[f"D{row}"].font = Font(bold=True)
        ws[f"E{row}"].font = Font(bold=True)
        ws[f"F{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = border
        ws[f"B{row}"].border = border
        ws[f"C{row}"].border = border
        ws[f"D{row}"].border = border
        ws[f"E{row}"].border = border
        ws[f"F{row}"].border = border

    for order in queryset:
        for index, item in enumerate(order.items.all(), start=1):
            # Calculate the total in UZS
            count += 1
            item_total = float(item.price_uzs) * float(item.qty)
            total_sum += int(item_total)

            # Format price and total with thousands separators
            item_total_formatted = "{:,.0f}".format(item_total).replace(",", " ")
            # Prepare the data to append
            case = int(round(float(item.qty))) - int(float(item.set_amount) * float(item.product_in_set))
            item_data = [
                count,
                item.product_name,
                item.product_in_set,
                f"{item.set_amount} Коробка {case} шт.",
                item.qty,
                item_total_formatted,
            ]

            ws.append(item_data)

            for col_num in range(1, len(item_data) + 1):
                cell = ws.cell(row=ws.max_row, column=col_num)
                cell.alignment = Alignment(horizontal="right")
                cell.border = border

            total_count += int(float(item.qty))

        items_count = order.items.all().count()
        row_count += items_count

    ws[f"A{count + 4}"] = ""
    ws[f"B{count + 4}"] = ""
    ws[f"C{count + 4}"] = ""
    ws[f"D{count + 4}"] = "ИТОГО"
    ws[f"E{count + 4}"] = total_count
    ws[f"D{count + 4}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"E{count + 4}"].font = Font(bold=True)

    ws[f"E{count + 4}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"F{count + 4}"].alignment = Alignment(horizontal="right", vertical="center")

    ws[f"F{count + 4}"] = "{:,.0f}".format(total_sum).replace(",", " ")
    ws[f"F{count + 4}"].number_format = '# ##0'

    for row in ws[f"A{count + 4}:F{count + 4}"]:
        for cell in row:
            cell.fill = fill
            cell.font = Font(bold=True)

    # Set up the HTTP response with the generated Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="orders.xlsx"'
    wb.save(response)
    return response


export_invoice_total_amount.short_description = "Накладная общая сумма (Excel)"

@admin.register(ArchiveOrder)
class ActiveOrderAdmin(admin.ModelAdmin):
    list_display = ("id", "user", "status", "comment", "get_total_cost", "location_path")
    list_per_page = 20
    inlines = (OrderItemTabularInline,)
    fields = (
        "user", "status", "payment_status", "payment_type", "comment", "is_rop_confirm", "is_accountant_confirm", "is_director_confirm",
        "is_storekeeper_confirm", "is_rop_cancel", "is_accountant_cancel", "is_director_cancel", "is_storekeeper_cancel")
    list_filter = ("user", "agent", "status", "payment_status", "user__territory",)
    search_fields = ("id",)
    date_hierarchy = "created_at"
    resource_class = OrderResource
    skip_export_form = True

    actions = ['generate_multiple_pdfs', 'generate_pdf2', export_orders_to_excel, export_invoice_total_amount]

    def generate_multiple_pdfs(self, request, queryset):
        selected_ids = queryset.values_list('id', flat=True)

        if len(selected_ids) == 1:
            return HttpResponseRedirect(f'/pdf/{selected_ids[0]}/')
        else:
            ids = ','.join(str(pk) for pk in selected_ids)
            return HttpResponseRedirect(f'/generate-multiple-pdfs/?ids={ids}')

    generate_multiple_pdfs.short_description = "Накладная для заказа (PDF)"

    def generate_pdf2(self, request, queryset):
        selected_ids = queryset.values_list('id', flat=True)
        ids = ','.join(str(pk) for pk in selected_ids)
        return HttpResponseRedirect(f'/pdf/?orders={ids}')

    generate_pdf2.short_description = "Накладная общая сумма (PDF)"

    def configure_ids(self, request, queryset):
        for order in queryset:
            for product in order.items.all():
                p = Product.objects.filter(title__icontains=product.product_name).last()
                product.product_id = p.pk if p else 0

    def get_readonly_fields(self, request, obj=None):
        if request.user.username == "admin" and request.user.is_superuser:
            return []
               
        return [field.name for field in self.model._meta.fields]


    def get_fields(self, request: HttpRequest, obj=None):
        if request.user.role == "storekeeper":
            return ("status", "user", "comment", "is_storekeeper_confirm")
        return super().get_fields(request, obj)

    def get_list_display(self, request):
        if request.user.role == "rop":
            return (
                "id", "user", "get_agent_name", "get_agent_territory", "status", "get_total_cost", "payment_status",
                "payment_type", "comment", "get_location", "created_at",
                "get_rop_approve_time", "get_accountant_approve_time", "get_director_approve_time",
                "get_storekeeper_approve_time",
            )

        if request.user.role == "accountant":
            return (
                "id", "user", "get_agent_name", "get_agent_territory", "status", "get_total_cost", "payment_status",
                "payment_type", "comment", "get_location", "created_at",
                "get_rop_approve_time", "get_accountant_approve_time", "get_director_approve_time",
                "get_storekeeper_approve_time",
            )

        if request.user.role == "director":
            return (
                "id", "user", "get_agent_name", "get_agent_territory", "status", "get_total_cost", "payment_status",
                "payment_type", "comment", "get_location", "created_at",
                "get_rop_approve_time", "get_accountant_approve_time", "get_director_approve_time",
                "get_storekeeper_approve_time",
            )

        if request.user.role == "storekeeper":
            return (
                "id", "user", "get_agent_name", "get_agent_territory", "status", "get_total_cost", "comment",
                "get_location", "created_at", "get_accountant_approve_time",
                "get_rop_approve_time", "get_director_approve_time", "get_storekeeper_approve_time",
            )

        return (
            "id", "user", "get_agent_name", "get_agent_territory", "status", "get_total_cost", "comment",
            "payment_status", "payment_type", "get_location",
            "created_at", "get_rop_approve_time", "get_accountant_approve_time", "get_director_approve_time",
            "get_storekeeper_approve_time",
        )

    def get_agent_territory(self, obj):
        if obj.agent and obj.agent.territory.exists():
            return ", ".join([area.name for area in obj.agent.territory.all()])
        return None

    get_agent_territory.short_description = "Территория"

    def get_agent_name(self, obj):
        if obj.agent:
            return obj.agent.get_full_name()
        return None

    get_agent_name.short_description = "Агент"

    def get_rop_approve_time(self, obj):
        if obj.is_rop_confirm and obj.rop_approve_time:
            return format_html(
                f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="green"/><path fill="none" stroke="white" stroke-width="2" d="M6 12l4 4l8-8" /></svg> Подтвержденный <br>{obj.rop_approve_time.strftime("%d.%m.%Y %H:%M:%S")}' if obj.is_rop_confirm else '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="red"/><path fill="none" stroke="white" stroke-width="2" d="M6 6l12 12M6 18L18 6" /></svg>')

    get_rop_approve_time.short_description = "Руководитель отдел продаж"

    def get_accountant_approve_time(self, obj):
        if obj.is_accountant_confirm and obj.accountant_approve_time:
            return format_html(
                f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="green"/><path fill="none" stroke="white" stroke-width="2" d="M6 12l4 4l8-8" /></svg> Подтвержденный <br>{obj.accountant_approve_time.strftime("%d.%m.%Y %H:%M:%S")}' if obj.is_accountant_confirm else '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="red"/><path fill="none" stroke="white" stroke-width="2" d="M6 6l12 12M6 18L18 6" /></svg>')

    get_accountant_approve_time.short_description = "Бухгалтер"

    def get_director_approve_time(self, obj):
        if obj.is_director_confirm and obj.director_approve_time:
            return format_html(
                f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="green"/><path fill="none" stroke="white" stroke-width="2" d="M6 12l4 4l8-8" /></svg> Подтвержденный <br>{obj.director_approve_time.strftime("%d.%m.%Y %H:%M:%S")}' if obj.is_director_confirm else '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="red"/><path fill="none" stroke="white" stroke-width="2" d="M6 6l12 12M6 18L18 6" /></svg>')

    get_director_approve_time.short_description = "Директор"

    def get_storekeeper_approve_time(self, obj):
        if obj.is_storekeeper_confirm and obj.storekeeper_approve_time:
            return format_html(
                f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="green"/><path fill="none" stroke="white" stroke-width="2" d="M6 12l4 4l8-8" /></svg> Подтвержденный <br>{obj.storekeeper_approve_time.strftime("%d.%m.%Y %H:%M:%S")}' if obj.is_storekeeper_confirm else '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="red"/><path fill="none" stroke="white" stroke-width="2" d="M6 6l12 12M6 18L18 6" /></svg>')

    get_storekeeper_approve_time.short_description = "Кладовщик"

    def formfield_for_choice_field(self, db_field, request, **kwargs):
        if db_field.name == 'status':
            if request.user.role == "accountant":
                kwargs['choices'] = Order.AccountantStatus.choices
            elif request.user.role == "director":
                kwargs['choices'] = Order.DirectorStatus.choices
            else:
                kwargs['choices'] = Order.StoreKeeperStatus.choices

        return super().formfield_for_choice_field(db_field, request, **kwargs)

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        queryset = queryset.annotate(
            total_price=Sum(F('items__price_uzs') * F('items__qty'), output_field=FloatField())
        )
        
        queryset = queryset.exclude(
            Q(is_rop_cancel=True) | 
            Q(is_accountant_cancel=True) |
            Q(is_director_cancel=True) |
            Q(is_storekeeper_cancel=True)
        )

        queryset = queryset.filter(
            is_rop_confirm=True,
            is_director_confirm=True,
            is_storekeeper_confirm=True,
            is_accountant_confirm=True
        )

        return queryset

    def has_delete_permission(self, request, obj=None):
        return False

    def has_add_permission(self, request):
        return False

    def get_total_cost(self, obj):
        return f"{(obj.total_price):,}" if obj.total_price else 0

    get_total_cost.short_description = "Общая сумма"

    def get_location(self, obj):
        if obj.location_path:
            return format_html(
                f"<a href='{obj.location_path}'>Посмотреть место нахождения</a>"
            )

    get_location.short_description = "Место нахождения"

    list_display_links = (
        "id", "user", "status", "get_total_cost", "payment_status", "payment_type", "get_location", "created_at",
        "get_rop_approve_time", "get_accountant_approve_time", "get_director_approve_time", "get_storekeeper_approve_time")


@admin.register(Order)
class OrderAdmin(ImportExportModelAdmin):
    list_display = ("id", "user", "status", "comment", "get_total_cost", "location_path")
    list_per_page = 20
    inlines = (OrderItemTabularInline,)
    fields = (
    "user", "status", "payment_status", "payment_type", "comment", "is_rop_confirm", "is_accountant_confirm", "is_director_confirm",
    "is_storekeeper_confirm", "is_rop_cancel", "is_accountant_cancel", "is_director_cancel", "is_storekeeper_cancel")
    list_filter = ("user", "agent", "status", "payment_status", "user__territory",)
    search_fields = ("id",)
    date_hierarchy = "created_at"
    resource_class = OrderResource
    skip_export_form = True
    

    actions = ['generate_multiple_pdfs', 'generate_pdf2', export_orders_to_excel, export_invoice_total_amount]

    def generate_multiple_pdfs(self, request, queryset):
        selected_ids = queryset.values_list('id', flat=True)

        if len(selected_ids) == 1:
            return HttpResponseRedirect(f'/pdf/{selected_ids[0]}/')
        else:
            ids = ','.join(str(pk) for pk in selected_ids)
            return HttpResponseRedirect(f'/generate-multiple-pdfs/?ids={ids}')

    generate_multiple_pdfs.short_description = "Накладная для заказа (PDF)"

    def generate_pdf2(self, request, queryset):
        selected_ids = queryset.values_list('id', flat=True)
        ids = ','.join(str(pk) for pk in selected_ids)
        return HttpResponseRedirect(f'/pdf/?orders={ids}')

    generate_pdf2.short_description = "Накладная общая сумма (PDF)"

    def configure_ids(self, request, queryset):
        for order in queryset:
            for product in order.items.all():
                p = Product.objects.filter(title__icontains=product.product_name).last()
                product.product_id = p.pk if p else 0

    def get_readonly_fields(self, request, obj=None):
        if request.user and request.user.username == "admin" and request.user.is_superuser:
            return []

        if obj and obj.status == "cancelled":
            return [field.name for field in self.model._meta.fields]

        if request.user.role == "rop" and obj.is_rop_confirm and obj.rop_approve_time:
            if timezone.now() <= obj.rop_approve_time + timedelta(hours=1):
                return [
                    field.name for field in self.model._meta.fields
                    if field.name not in {"is_rop_confirm", "is_rop_cancel"}
                ]
        
        if request.user.role == "accountant" and obj.is_accountant_confirm and obj.accountant_approve_time:
            if timezone.now() <= obj.accountant_approve_time + timedelta(hours=1):
                return [
                    field.name for field in self.model._meta.fields
                    if field.name not in {"is_accountant_confirm", "is_accountant_cancel"}
                ]
        
        if request.user.role == "director" and obj.is_director_confirm and obj.director_approve_time:
            if timezone.now() <= obj.director_approve_time + timedelta(hours=1):
                return [
                    field.name for field in self.model._meta.fields
                    if field.name not in {"is_director_confirm", "is_director_cancel"}
                ]
        
        if request.user.role == "storekeeper" and obj.is_storekeeper_confirm and obj.storekeeper_approve_time:
            if timezone.now() <= obj.storekeeper_approve_time + timedelta(hours=1):
                return [
                    field.name for field in self.model._meta.fields
                    if field.name not in {"is_storekeeper_confirm", "is_storekeeper_cancel"}
                ]

        if request.user.role == "rop":
            if obj.is_rop_confirm or obj.is_rop_cancel:
                return [field.name for field in self.model._meta.fields]

            return [
                field.name for field in self.model._meta.fields
                if field.name not in {"is_rop_confirm", "is_rop_cancel"}
            ]

        if request.user.role == "accountant":
            if obj.is_rop_confirm and not obj.is_accountant_confirm:
                return [
                    field.name for field in self.model._meta.fields
                    if field.name not in {"is_accountant_confirm", "is_accountant_cancel"}
                ]

            if obj.is_accountant_confirm or obj.is_accountant_cancel or obj.is_rop_cancel:
                return [field.name for field in self.model._meta.fields]

        if request.user.role == "director":
            if obj.is_accountant_confirm and not obj.is_director_confirm:
                return [
                    field.name for field in self.model._meta.fields
                    if field.name not in {"is_director_confirm", "is_director_cancel"}
                ]
                
            if obj.is_director_confirm or obj.is_director_cancel or obj.is_accountant_cancel:
                return [field.name for field in self.model._meta.fields]
    
        if request.user.role == "storekeeper":
            if obj.is_director_confirm and not obj.is_storekeeper_confirm:
                return [
                    field.name for field in self.model._meta.fields
                    if field.name not in {"is_storekeeper_confirm", "is_storekeeper_cancel"}
                ]
            
            if obj.is_storekeeper_confirm or obj.is_storekeeper_cancel or obj.is_director_cancel:
                return [field.name for field in self.model._meta.fields]
            
        return [field.name for field in self.model._meta.fields]


    def get_fields(self, request: HttpRequest, obj=None):
        if request.user.role == "storekeeper":
            return ("status", "user", "comment", "is_storekeeper_confirm")
        return super().get_fields(request, obj)

    def get_list_display(self, request):
        if request.user.role == "rop":
            return (
                "id", "user", "get_agent_name", "get_agent_territory", "status", "get_total_cost", "payment_status", "payment_type", "comment", "get_location", "created_at",
                "get_rop_approve_time", "get_accountant_approve_time", "get_director_approve_time", "get_storekeeper_approve_time",
                "get_rop_cancel_time", "get_accountant_cancel_time", "get_director_cancel_time", "get_storekeeper_cancel_time"
            )

        if request.user.role == "accountant":
            return (
                "id", "user", "get_agent_name", "get_agent_territory", "status", "get_total_cost", "payment_status", "payment_type", "comment", "get_location", "created_at",
                "get_rop_approve_time", "get_accountant_approve_time", "get_director_approve_time", "get_storekeeper_approve_time",
                "get_rop_cancel_time", "get_accountant_cancel_time", "get_director_cancel_time", "get_storekeeper_cancel_time"
            )

        if request.user.role == "director":
            return (
                "id", "user", "get_agent_name", "get_agent_territory", "status", "get_total_cost", "payment_status", "payment_type", "comment", "get_location", "created_at",
                "get_rop_approve_time", "get_accountant_approve_time", "get_director_approve_time", "get_storekeeper_approve_time",
                "get_rop_cancel_time", "get_accountant_cancel_time", "get_director_cancel_time", "get_storekeeper_cancel_time",
            )

        if request.user.role == "storekeeper":
            return (
                "id", "user", "get_agent_name", "get_agent_territory", "status", "get_total_cost", "comment", "get_location", "created_at", "get_accountant_approve_time",
                "get_rop_approve_time", "get_director_approve_time", "get_storekeeper_approve_time",
                "get_rop_cancel_time", "get_accountant_cancel_time", "get_director_cancel_time", "get_storekeeper_cancel_time",
            )

        return (
            "id", "user", "get_agent_name", "get_agent_territory", "status", "get_total_cost", "comment", "payment_status", "payment_type", "get_location",
            "created_at", "get_rop_approve_time", "get_accountant_approve_time", "get_director_approve_time",
            "get_storekeeper_approve_time", "get_rop_cancel_time", "get_accountant_cancel_time", "get_director_cancel_time", "get_storekeeper_cancel_time"
        )

    def get_agent_territory(self, obj):
        if obj.agent and obj.agent.territory.exists():
            return ", ".join([area.name for area in obj.agent.territory.all()])
        return None

    get_agent_territory.short_description = "Территория"

    def get_agent_name(self, obj):
        if obj.agent:
            return obj.agent.get_full_name()
        return None

    get_agent_name.short_description = "Агент"

    def get_rop_cancel_time(self, obj):
        return format_html(
            f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="green"/><path fill="none" stroke="white" stroke-width="2" d="M6 12l4 4l8-8" /></svg> Отказанный <br>{obj.rop_cancel_time.strftime("%d.%m.%Y %H:%M:%S")}' if obj.rop_cancel_time else '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="red"/><path fill="none" stroke="white" stroke-width="2" d="M6 6l12 12M6 18L18 6" /></svg>')

    get_rop_cancel_time.short_description = "Отказ руководителем отдела продаж"

    def get_accountant_cancel_time(self, obj):
        return format_html(
            f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="green"/><path fill="none" stroke="white" stroke-width="2" d="M6 12l4 4l8-8" /></svg> Отказанный <br>{obj.accountant_cancel_time.strftime("%d.%m.%Y %H:%M:%S")}' if obj.accountant_cancel_time else '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="red"/><path fill="none" stroke="white" stroke-width="2" d="M6 6l12 12M6 18L18 6" /></svg>')

    get_accountant_cancel_time.short_description = "Отказ бухгалтером"

    def get_director_cancel_time(self, obj):
        return format_html(
            f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="green"/><path fill="none" stroke="white" stroke-width="2" d="M6 12l4 4l8-8" /></svg> Отказанно <br>{obj.director_cancel_time.strftime("%d.%m.%Y %H:%M:%S")}' if obj.director_cancel_time else '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="red"/><path fill="none" stroke="white" stroke-width="2" d="M6 6l12 12M6 18L18 6" /></svg>')

    get_director_cancel_time.short_description = "Отказ директором"

    def get_storekeeper_cancel_time(self, obj):
        return format_html(
            f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="green"/><path fill="none" stroke="white" stroke-width="2" d="M6 12l4 4l8-8" /></svg> Отказанно <br>{obj.storekeeper_cancel_time.strftime("%d.%m.%Y %H:%M:%S")}' if obj.storekeeper_cancel_time else '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="red"/><path fill="none" stroke="white" stroke-width="2" d="M6 6l12 12M6 18L18 6" /></svg>')

    get_storekeeper_cancel_time.short_description = "Отказ кладовщиком"

    def get_rop_approve_time(self, obj):
        if obj and obj.status == "cancelled" and obj.rop_approve_time:
            return format_html(
                '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="red"/><path fill="none" stroke="white" stroke-width="2" d="M6 6l12 12M6 18L18 6" /></svg> Отменено')
        return format_html(
            f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="green"/><path fill="none" stroke="white" stroke-width="2" d="M6 12l4 4l8-8" /></svg> Подтвержденный <br>{obj.rop_approve_time.strftime("%d.%m.%Y %H:%M:%S")}' if obj.rop_approve_time else '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="red"/><path fill="none" stroke="white" stroke-width="2" d="M6 6l12 12M6 18L18 6" /></svg>')

    get_rop_approve_time.short_description = "Руководитель отдел продаж"

    def get_accountant_approve_time(self, obj):
        if obj and obj.status == "cancelled" and obj.accountant_approve_time:
            return format_html(
                '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="red"/><path fill="none" stroke="white" stroke-width="2" d="M6 6l12 12M6 18L18 6" /></svg> Отменено')
        return format_html(
            f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="green"/><path fill="none" stroke="white" stroke-width="2" d="M6 12l4 4l8-8" /></svg> Подтвержденный <br>{obj.accountant_approve_time.strftime("%d.%m.%Y %H:%M:%S")}' if obj.accountant_approve_time else '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="red"/><path fill="none" stroke="white" stroke-width="2" d="M6 6l12 12M6 18L18 6" /></svg>')

    get_accountant_approve_time.short_description = "Бухгалтер"

    def get_director_approve_time(self, obj):
        if obj and obj.status == "cancelled" and obj.director_approve_time:
            return format_html(
                '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="red"/><path fill="none" stroke="white" stroke-width="2" d="M6 6l12 12M6 18L18 6" /></svg> Отменено')
        return format_html(
            f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="green"/><path fill="none" stroke="white" stroke-width="2" d="M6 12l4 4l8-8" /></svg> Подтвержденный <br>{obj.director_approve_time.strftime("%d.%m.%Y %H:%M:%S")}' if obj.director_approve_time else '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="red"/><path fill="none" stroke="white" stroke-width="2" d="M6 6l12 12M6 18L18 6" /></svg>')

    get_director_approve_time.short_description = "Директор"

    def get_storekeeper_approve_time(self, obj):
        return format_html(
            f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="green"/><path fill="none" stroke="white" stroke-width="2" d="M6 12l4 4l8-8" /></svg> Подтвержденный <br>{obj.storekeeper_approve_time.strftime("%d.%m.%Y %H:%M:%S")}' if obj.storekeeper_approve_time else '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 24 24" width="24" height="24"><circle cx="12" cy="12" r="12" fill="red"/><path fill="none" stroke="white" stroke-width="2" d="M6 6l12 12M6 18L18 6" /></svg>')

    get_storekeeper_approve_time.short_description = "Кладовщик"

    def formfield_for_choice_field(self, db_field, request, **kwargs):
        if db_field.name == 'status':
            if request.user.role == "accountant":
                kwargs['choices'] = Order.AccountantStatus.choices
            elif request.user.role == "director":
                kwargs['choices'] = Order.DirectorStatus.choices
            else:
                kwargs['choices'] = Order.StoreKeeperStatus.choices

        return super().formfield_for_choice_field(db_field, request, **kwargs)

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        queryset = queryset.annotate(
            total_price=Sum(F('items__price_uzs') * F('items__qty'), output_field=FloatField())
        )

        if request.user.role == "director":
            queryset = queryset.exclude(status="pending")

        if request.user.role == "storekeeper":
            queryset = queryset.exclude(Q(status="pending") | Q(status="approved_by_account"))
        return queryset

    def has_delete_permission(self, request, obj=None):
        return False

    def has_add_permission(self, request):
        return False

    def get_total_cost(self, obj):
        return f"{(obj.total_price):,}" if obj.total_price else 0

    get_total_cost.short_description = "Общая сумма"

    def get_location(self, obj):
        if obj.location_path:
            return format_html(
                f"<a href='{obj.location_path}'>Посмотреть место нахождения</a>"
            )

    get_location.short_description = "Место нахождения"

    list_display_links = (
    "id", "user", "status", "get_total_cost", "payment_status", "payment_type", "get_location", "created_at",
    "get_rop_approve_time", "get_accountant_approve_time", "get_director_approve_time", "get_storekeeper_approve_time",
    "get_rop_cancel_time", "get_accountant_cancel_time", "get_director_cancel_time", "get_storekeeper_cancel_time",)


@admin.register(Area)
class AreaAdmin(admin.ModelAdmin):
    list_display = ("id", "name",)
    search_fields = ("name",)
    list_display_links = list_display

    def get_search_results(self, request, queryset, search_term):
        if search_term:
            search_term = search_term.lower()  # Ensure the search term is lowercase
            queries = [
                Q(**{f"{field}__icontains": search_term}) |
                Q(**{f"{field}__iregex": f"(?i){search_term}"})  # Using regex for case-insensitive match
                for field in self.search_fields
            ]
            queryset = queryset.filter(*queries)
        return queryset, False
