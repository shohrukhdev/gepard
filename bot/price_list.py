import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Product

def export_products_to_excel(request):
    # Create an Excel workbook and worksheet
    cat = request.GET.get('cat')
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Products"

    # Define the header for the Excel file
    headers = [
        "ID", "Название продукта", "Цена (сум)"
    ]

    # Write the headers to the worksheet
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        worksheet[f"{col_letter}1"] = header

    # Query the products and write their data to the worksheet
    products = Product.objects.all()
    products = products.extra(
                    select={
                        'price_uzs': f'price_uzs_{cat}',
                        'price_usd': f'price_usd_{cat}'
                    }
                )
    for row_num, product in enumerate(products, 2):
        worksheet[f"A{row_num}"] = product.id
        worksheet[f"B{row_num}"] = product.title
        worksheet[f"C{row_num}"] = product.price_uzs
        

    # Prepare the response to download the Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="products.xlsx"'
    
    # Save the workbook to the response
    workbook.save(response)
    return response
